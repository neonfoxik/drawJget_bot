import telebot
import os
from dotenv import load_dotenv
import random
import datetime
import openpyxl
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
import string

load_dotenv()

BOT_TOKEN = os.getenv('BOT_TOKEN')

if not BOT_TOKEN:
    print("❌ Ошибка: BOT_TOKEN не найден в переменных окружения!")
    print("Создайте файл .env с токеном бота")
    exit(1)

CHANNEL_USERNAME = os.getenv('CHANNEL_USERNAME', 'official_jget')

bot = telebot.TeleBot(BOT_TOKEN)

user_states = {}

def generate_participant_code():
    """Генерирует код участника"""
    characters = string.ascii_uppercase + string.digits
    return ''.join(random.choice(characters) for _ in range(6))

# Функция для проверки подписки на канал
def check_subscription(user_id, channel_username):
    """Проверяет, подписан ли пользователь на канал"""
    try:
        channel_variants = [
            channel_username, 
            f"@{channel_username}", 
            f"https://t.me/{channel_username}", 
        ]
        
        for channel in channel_variants:
            try:
                member = bot.get_chat_member(channel, user_id)
                
                if member.status in ['member', 'administrator', 'creator']:
                    return True
            except Exception as e:
                print(f"DEBUG: Ошибка при проверке канала {channel}: {e}")
                continue
        
        return False
    except Exception as e:
        print(f"DEBUG: Общая ошибка при проверке подписки: {e}")
        return False

def find_existing_participant(user_id):
    """Ищет существующего участника в Excel таблице"""
    filename = 'participants.xlsx'
    
    try:
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            
            for row in range(2, sheet.max_row + 1):  
                if sheet.cell(row=row, column=1).value == user_id:
                    return {
                        'username': sheet.cell(row=row, column=2).value,
                        'parent_fio': sheet.cell(row=row, column=3).value,
                        'child_school': sheet.cell(row=row, column=4).value,
                        'child_class': sheet.cell(row=row, column=5).value,
                        'phone_number': sheet.cell(row=row, column=6).value,
                        'participant_code': sheet.cell(row=row, column=7).value,
                        'registration_date': sheet.cell(row=row, column=8).value
                    }
            
            return None  
            
        except FileNotFoundError:
            return None 
            
    except Exception as e:
        print(f"Ошибка при поиске участника в Excel: {e}")
        return None

def save_participant_to_excel(user_id, username, parent_fio, child_school, child_class, phone_number, participant_code):
    """Сохраняет данные участника в Excel файл"""
    filename = 'participants.xlsx'
    
    try:
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['ID пользователя', 'Username Telegram', 'ФИО родителя', 'Школа ребенка', 'Класс ребенка', 'Номер телефона', 'Код участника', 'Дата регистрации'])
        
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([user_id, username, parent_fio, child_school, child_class, phone_number, participant_code, current_time])
        
        workbook.save(filename)
        return True
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {e}")
        return False

def create_main_keyboard():
    """Создает основную клавиатуру с кнопками"""
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("🔍 Подписаться на канал", callback_data="check_subscription")
    )
    return keyboard

def create_phone_keyboard():
    """Создает клавиатуру с кнопкой отправки номера телефона"""
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(KeyboardButton("📱 Отправить мой номер", request_contact=True))
    return keyboard

def create_info_keyboard():
    """Создает клавиатуру для сбора информации о родителе и ребенке"""
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(KeyboardButton("📝 Ввести информацию"))
    return keyboard

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def send_welcome(message):
    user_id = message.from_user.id
    username = message.from_user.username
    first_name = message.from_user.first_name
    
    # Проверяем, есть ли уже участник в таблице
    existing_participant = find_existing_participant(user_id)
    
    if existing_participant:
        # Участник уже зарегистрирован - показываем его данные
        existing_text = f"""
🎉 **Добро пожаловать обратно!**

✅ **Вы уже участник розыгрыша!**

📋 **Ваши данные:**
👤 ФИО родителя: {existing_participant['parent_fio']}
🏫 Школа ребенка: {existing_participant['child_school']}
📚 Класс ребенка: {existing_participant['child_class']}
📱 Телефон: {existing_participant['phone_number']}
🎫 **Код участника: `{existing_participant['participant_code']}`**
📅 Дата регистрации: {existing_participant['registration_date']}

⏳ **Ждите результаты розыгрыша!**
Удачи! 🍀
        """
        
        # Инициализируем состояние как уже зарегистрированного
        user_states[user_id] = {
            'username': username,
            'first_name': first_name,
            'subscribed': True,  # Считаем что подписан
            'phone_provided': True,  # Считаем что номер предоставлен
            'phone_number': existing_participant['phone_number'],
            'parent_fio': existing_participant['parent_fio'],
            'child_school': existing_participant['child_school'],
            'child_class': existing_participant['child_class']
        }
        
        bot.reply_to(message, existing_text, parse_mode='Markdown')
        
    else:
        # Новый участник - начинаем регистрацию
        # Инициализируем состояние пользователя
        user_states[user_id] = {
            'username': username,
            'first_name': first_name,
            'subscribed': False,
            'phone_provided': False,
            'phone_number': None,
            'parent_fio': None,
            'child_school': None,
            'child_class': None,
            'current_step': 'subscription'  # Текущий шаг регистрации
        }
        
        welcome_text = f"""
Привет! 👋

Я бот для проведения розыгрыша от клуба робототехники J-Get!

Чтобы стать участником розыгрыша:
1️⃣ Подпишитесь на наш канал @{CHANNEL_USERNAME}
2️⃣ Предоставьте свой номер телефона
3️⃣ Введите информацию о родителе и ребенке

Нажмите кнопки ниже для выполнения требований:
        """
        
        bot.reply_to(message, welcome_text, reply_markup=create_main_keyboard())

# Обработчик callback запросов от inline кнопок
@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    user_id = call.from_user.id
    
    if call.data == "check_subscription":
        # Проверяем подписку на канал
        if check_subscription(user_id, CHANNEL_USERNAME):
            # Убеждаемся, что состояние пользователя существует
            if user_id not in user_states:
                user_states[user_id] = {
                    'username': call.from_user.username,
                    'first_name': call.from_user.first_name,
                    'subscribed': False,
                    'phone_provided': False,
                    'phone_number': None
                }
            
            user_states[user_id]['subscribed'] = True
            bot.answer_callback_query(call.id, "✅ Вы подписаны на канал!")
            
            # После успешной подписки показываем только кнопку для номера телефона
            user_states[user_id]['current_step'] = 'phone'
            new_text = f"""
✅ **Вы подписаны на канал @{CHANNEL_USERNAME}!**

📋 **Статус требований:**
🔍 Подписка на канал: ✅ Выполнено
📱 Номер телефона: ❌ Не предоставлен
📝 Информация о родителе и ребенке: ❌ Не введена

**Следующий шаг:** Предоставьте свой номер телефона
            """
            
            keyboard = InlineKeyboardMarkup().add(
                InlineKeyboardButton("📱 Предоставить номер", callback_data="provide_phone")
            )
            
            bot.edit_message_text(
                new_text,
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                reply_markup=keyboard
            )
        else:
            # Пользователь не подписан - показываем полупрозрачное сообщение
            bot.answer_callback_query(
                call.id, 
                f"❌ Вы не подписаны на канал @{CHANNEL_USERNAME}!", 
                show_alert=True
            )
            
           
    elif call.data == "provide_phone":
        # Убеждаемся, что состояние пользователя существует
        if user_id not in user_states:
            user_states[user_id] = {
                'username': call.from_user.username,
                'first_name': call.from_user.first_name,
                'subscribed': False,
                'phone_provided': False,
                'phone_number': None
            }
        
        # Проверяем, подписан ли пользователь на канал
        if not user_states[user_id]['subscribed']:
            bot.answer_callback_query(
                call.id, 
                "❌ Сначала подпишитесь на канал!", 
                show_alert=True
            )
            return
        
        # Запрашиваем номер телефона
        bot.answer_callback_query(call.id, "Введите номер телефона")
        bot.send_message(
            call.message.chat.id,
            "Пожалуйста, нажмите кнопку ниже, чтобы поделиться своим номером телефона:",
            reply_markup=create_phone_keyboard()
        )
    

    


# Обработчик получения контакта (номера телефона)
@bot.message_handler(content_types=['contact'])
def handle_contact(message):
    user_id = message.from_user.id
    
    if user_id not in user_states:
        bot.reply_to(message, "Пожалуйста, начните с команды /start")
        return
    
    # Получаем номер телефона
    phone_number = message.contact.phone_number
    user_states[user_id]['phone_number'] = phone_number
    user_states[user_id]['phone_provided'] = True
    
    # После получения номера телефона начинаем сбор информации пошагово
    user_states[user_id]['current_step'] = 'parent_fio'
    
    success_text = f"""
✅ **Номер телефона получен!**

📋 **Статус требований:**
🔍 Подписка на канал: ✅ Выполнено
📱 Номер телефона: ✅ Предоставлен
📝 Информация о родителе и ребенке: ❌ Не введена

**Следующий шаг:** Введите ФИО родителя

Пожалуйста, напишите ваше полное имя (например: Иванов Иван Иванович)
    """
    
    bot.reply_to(message, success_text, parse_mode='Markdown')

@bot.message_handler(func=lambda message: True)
def echo_all(message):
    user_id = message.from_user.id
    user_text = message.text.lower()
    
    # Проверяем, есть ли пользователь в состоянии
    if user_id not in user_states:
        bot.reply_to(message, "Пожалуйста, начните с команды /start")
        return
    
    # Простые ответы на ключевые слова
    if 'привет' in user_text or 'здравствуй' in user_text:
        bot.reply_to(message, f"Привет! Рад тебя видеть! 😊")
    elif 'как дела' in user_text or 'как ты' in user_text:
        bot.reply_to(message, "У меня всё отлично! А у тебя как дела? 😄")
    elif 'спасибо' in user_text or 'благодарю' in user_text:
        bot.reply_to(message, "Пожалуйста! Рад быть полезным! 😊")
    elif 'пока' in user_text or 'до свидания' in user_text:
        bot.reply_to(message, "До свидания! Буду ждать нашей следующей встречи! 👋")
    elif 'кто ты' in user_text or 'что ты умеешь' in user_text:
        bot.reply_to(message, "Я бот для проведения розыгрыша от клуба робототехники J-Get! Умею регистрировать участников и вести простые диалоги. Попробуйте команду /help для списка возможностей! 🤖")
    else:
        # Пытаемся обработать пошаговый сбор информации
        if handle_registration_steps(message):
            return
        else:
            # Если не распознали ключевые слова и не обрабатывали регистрацию, отправляем эхо
            bot.reply_to(message, f"Вы написали: {message.text}")

def handle_registration_steps(message):
    """Обрабатывает пошаговый сбор информации для регистрации"""
    user_id = message.from_user.id
    text = message.text.strip()
    
    if user_id not in user_states:
        return False
    
    current_step = user_states[user_id].get('current_step', 'subscription')
    
    if current_step == 'parent_fio':
        # Сохраняем ФИО родителя
        user_states[user_id]['parent_fio'] = text
        user_states[user_id]['current_step'] = 'child_school'
        
        next_text = f"""
✅ **ФИО родителя сохранено: {text}**

📋 **Статус требований:**
🔍 Подписка на канал: ✅ Выполнено
📱 Номер телефона: ✅ Предоставлен
👤 ФИО родителя: ✅ Введено
🏫 Школа ребенка: ❌ Не введена
📚 Класс ребенка: ❌ Не введен

**Следующий шаг:** Введите школу ребенка

Пожалуйста, напишите название школы (например: МБОУ СОШ №39)
        """
        bot.reply_to(message, next_text, parse_mode='Markdown')
        return True
        
    elif current_step == 'child_school':
        # Сохраняем школу ребенка
        user_states[user_id]['child_school'] = text
        user_states[user_id]['current_step'] = 'child_class'
        
        next_text = f"""
✅ **Школа ребенка сохранена: {text}**

📋 **Статус требований:**
🔍 Подписка на канал: ✅ Выполнено
📱 Номер телефона: ✅ Предоставлен
👤 ФИО родителя: ✅ Введено
🏫 Школа ребенка: ✅ Введена
📚 Класс ребенка: ❌ Не введен

**Следующий шаг:** Введите класс ребенка

Пожалуйста, напишите класс (например: 7А, 8Б, 11)
        """
        bot.reply_to(message, next_text, parse_mode='Markdown')
        return True
        
    elif current_step == 'child_class':
        # Сохраняем класс ребенка и завершаем регистрацию
        user_states[user_id]['child_class'] = text
        
        # Все требования выполнены - генерируем код и регистрируем
        participant_code = generate_participant_code()
        
        # Сохраняем в Excel
        if save_participant_to_excel(
            user_id, 
            user_states[user_id]['username'], 
            user_states[user_id]['parent_fio'],
            user_states[user_id]['child_school'],
            text,  # child_class
            user_states[user_id]['phone_number'], 
            participant_code
        ):
            success_text = f"""
🎉 **Поздравляем! Теперь вы участник розыгрыша!**

📋 **Ваши данные:**
👤 ФИО родителя: {user_states[user_id]['parent_fio']}
🏫 Школа ребенка: {user_states[user_id]['child_school']}
📚 Класс ребенка: {text}
📱 Телефон: {user_states[user_id]['phone_number']}
🎫 **Код участника: `{participant_code}`**

💾 Сохраните этот код! Он понадобится для участия в розыгрыше.

🍀 Удачи в розыгрыше!
            """
            bot.reply_to(message, success_text, parse_mode='Markdown')
            
            # Отправляем отдельное сообщение с кодом
            bot.send_message(
                message.chat.id,
                f"🎫 **Ваш код участника: `{participant_code}`**",
                parse_mode='Markdown'
            )
            
            # Сбрасываем шаг регистрации
            user_states[user_id]['current_step'] = 'completed'
            return True
        else:
            bot.reply_to(message, "❌ Произошла ошибка при регистрации. Попробуйте позже.")
            return True
    
    return False

# Обработчик стикеров
@bot.message_handler(content_types=['sticker'])
def handle_sticker(message):
    bot.reply_to(message, "Классный стикер! 😄")

# Обработчик фото
@bot.message_handler(content_types=['photo'])
def handle_photo(message):
    bot.reply_to(message, "Красивое фото! 📸")

# Обработчик голосовых сообщений
@bot.message_handler(content_types=['voice'])
def handle_voice(message):
    bot.reply_to(message, "Получил ваше голосовое сообщение! 🎤")

if __name__ == "__main__":
    bot.polling(none_stop=True, interval=0)
